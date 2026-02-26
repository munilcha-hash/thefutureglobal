"""
플랫폼별 RAW 데이터 파일(CSV/Excel)을 DB로 임포트하는 관리 커맨드.
파일명에서 플랫폼을 자동 감지하거나 --platform으로 지정 가능.

CSV는 Python csv 모듈(경량), Excel은 openpyxl(경량)만 사용 - pandas 미사용.
스트리밍 배치 처리로 메모리 사용 최소화 (Railway Docker 최적화).

Usage:
  python manage.py import_raw path/to/file.csv
  python manage.py import_raw path/to/file.xlsx --platform shopee
"""
import csv
import gc
import os
import re
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from sales.models import ShopifyOrder, TiktokOrder, ShopeeOrder, Qoo10Order

BATCH_SIZE = 300


def safe_decimal(val, default=0):
    if val is None or val == '' or val == '-':
        return Decimal(str(default)) if default is not None else None
    try:
        cleaned = str(val).replace(',', '').replace('$', '').replace('\t', '').strip()
        if not cleaned or cleaned == '-':
            return Decimal(str(default)) if default is not None else None
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return Decimal(str(default)) if default is not None else None


def safe_str(val, default=''):
    if val is None:
        return default
    return str(val).replace('\t', '').strip()


def safe_int(val, default=0):
    if val is None or val == '':
        return default
    try:
        return int(float(str(val).replace(',', '').replace('\t', '').strip()))
    except (ValueError, TypeError):
        return default


def safe_date(val):
    """다양한 날짜 형식 파싱"""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        if isinstance(val, datetime):
            return val.date()
        return val
    s = str(val).replace('\t', '').strip()
    if not s:
        return None
    formats = [
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d',
        '%m/%d/%Y %I:%M:%S %p',
        '%m/%d/%Y %H:%M:%S',
        '%m/%d/%Y',
        '%d-%m-%Y %H:%M',
        '%d-%m-%Y',
    ]
    clean = s.split('+')[0].strip()
    clean = re.sub(r'\s*-\d{4}$', '', clean)
    for fmt in formats:
        try:
            return datetime.strptime(clean, fmt).date()
        except ValueError:
            continue
    return None


def detect_platform(filename):
    """파일명에서 플랫폼 자동 감지"""
    fname = filename.lower()
    if 'orders_export' in fname or '쇼피파이' in fname or 'shopify' in fname:
        return 'shopify'
    if 'all order' in fname or 'all_order' in fname or '틱톡' in fname or 'tiktok' in fname:
        return 'tiktok'
    if 'shopee' in fname or 'shop-stats' in fname or '쇼피' in fname:
        return 'shopee'
    if 'qoo10' in fname or 'transaction' in fname or '큐텐' in fname:
        return 'qoo10'
    return None


def extract_date_from_filename(filename):
    """파일명에서 날짜 추출 (YYYYMMDD 패턴)"""
    matches = re.findall(r'(\d{8})', filename)
    if matches:
        try:
            return datetime.strptime(matches[0], '%Y%m%d').date()
        except ValueError:
            pass
    return None


def extract_brand_from_shopee_filename(filename):
    """Shopee 파일명에서 브랜드 추출"""
    match = re.search(r'_([a-zA-Z]+)\.\w+\.shopee', filename)
    if match:
        brand = match.group(1)
        brand_mapping = {
            'drblet': '닥터블릿', 'doctorblet': '닥터블릿',
            'eoa': 'EOA', 'nothingviral': '낫띵베럴',
            'nothingbetter': '낫띵베럴', 'tetracure': '테트라큐어', 'calo': 'Calo',
        }
        return brand_mapping.get(brand.lower(), brand)
    return ''


# ─── 스트리밍 헬퍼 ─────────────────────────────────────────────

def _csv_stream(file_path):
    """CSV를 한 행씩 yield하는 제너레이터 (메모리 최소화)"""
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        if reader.fieldnames:
            reader.fieldnames = [h.replace('\t', '').strip() for h in reader.fieldnames]
        for row in reader:
            yield row


def _collect_csv_dates(file_path, date_field, alt_field=None):
    """CSV에서 날짜만 수집 (set으로 중복 없이, 메모리 최소)"""
    dates = set()
    for row in _csv_stream(file_path):
        val = row.get(date_field) or (row.get(alt_field) if alt_field else None)
        d = safe_date(val)
        if d:
            dates.add(d)
    return dates


def _flush_batch(model_class, batch):
    """배치를 DB에 저장하고 비움"""
    if batch:
        model_class.objects.bulk_create(batch, batch_size=BATCH_SIZE)
        n = len(batch)
        batch.clear()
        gc.collect()
        return n
    return 0


# ─── 커맨드 ─────────────────────────────────────────────

class Command(BaseCommand):
    help = '플랫폼별 RAW 데이터 파일(CSV/Excel)을 DB로 임포트'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='RAW 데이터 파일 경로')
        parser.add_argument('--platform', type=str, default=None,
                            choices=['shopify', 'tiktok', 'shopee', 'qoo10'],
                            help='플랫폼 (미지정시 파일명에서 자동 감지)')
        parser.add_argument('--clear-date', action='store_true',
                            help='해당 날짜 범위의 기존 데이터 삭제 후 임포트')
        parser.add_argument('--original-filename', type=str, default=None,
                            help='원본 파일명 (UUID 임시파일 사용 시 날짜/브랜드 추출용)')

    def handle(self, *args, **options):
        file_path = options['file_path']
        filename = options['original_filename'] or os.path.basename(file_path)

        platform = options['platform'] or detect_platform(filename)
        if not platform:
            raise CommandError(
                f'플랫폼을 감지할 수 없습니다: {filename}\n'
                f'--platform 옵션으로 지정해주세요 (shopify, tiktok, shopee, qoo10)'
            )

        # 파일 존재 확인
        if not os.path.exists(file_path):
            raise CommandError(f'파일을 찾을 수 없습니다: {file_path}')

        self.stdout.write(f"플랫폼: {platform.upper()} | 파일: {filename}")
        self.stdout.write(f"파일 크기: {os.path.getsize(file_path)} bytes")

        if platform == 'shopify':
            count = self._import_shopify_csv(file_path, options['clear_date'])
        elif platform == 'tiktok':
            count = self._import_tiktok_csv(file_path, options['clear_date'])
        elif platform == 'shopee':
            count = self._import_shopee_excel(file_path, filename, options['clear_date'])
        elif platform == 'qoo10':
            count = self._import_qoo10_excel(file_path, filename, options['clear_date'])
        else:
            count = 0

        if count == 0:
            raise CommandError(f'[{platform.upper()}] 임포트할 유효한 데이터가 없습니다.')

        gc.collect()
        self.stdout.write(self.style.SUCCESS(f"[{platform.upper()}] {count}건 임포트 완료!"))

    @transaction.atomic
    def _import_shopify_csv(self, file_path, clear_date):
        """Shopify orders_export CSV - 스트리밍 배치 임포트"""
        # Pass 1: 날짜 수집 (clear_date용, 메모리 최소)
        if clear_date:
            dates = _collect_csv_dates(file_path, 'Paid at', 'Created at')
            if dates:
                min_d, max_d = min(dates), max(dates)
                deleted = ShopifyOrder.objects.filter(
                    region='us', order_date__gte=min_d, order_date__lte=max_d
                ).delete()[0]
                self.stdout.write(f"  기존 데이터 {deleted}건 삭제 ({min_d} ~ {max_d})")
            del dates

        # Pass 2: 스트리밍 임포트
        batch = []
        total = 0
        for row in _csv_stream(file_path):
            order_date = safe_date(row.get('Paid at') or row.get('Created at'))
            if not order_date:
                continue

            brand = safe_str(row.get('Vendor', ''))
            name = safe_str(row.get('Name', ''))
            if not name and not brand:
                continue

            total_amt = safe_decimal(row.get('Total'), None)
            subtotal = safe_decimal(row.get('Subtotal'), None)

            batch.append(ShopifyOrder(
                region='us',
                brand=brand,
                final_amount=total_amt if total_amt is not None else subtotal,
                order_date=order_date,
                order_name=name,
                email=safe_str(row.get('Email', '')),
                financial_status=safe_str(row.get('Financial Status', '')),
                subtotal=subtotal,
                shipping_cost=safe_decimal(row.get('Shipping'), None),
                taxes=safe_decimal(row.get('Taxes'), None),
                total=total_amt,
                discount_code=safe_str(row.get('Discount Code', '')),
                discount_amount=safe_decimal(row.get('Discount Amount'), None),
                lineitem_quantity=safe_int(row.get('Lineitem quantity', 0)),
                lineitem_name=safe_str(row.get('Lineitem name', '')),
                lineitem_price=safe_decimal(row.get('Lineitem price'), None),
                lineitem_sku=safe_str(row.get('Lineitem sku', '')),
                shipping_city=safe_str(row.get('Shipping City', '')),
                shipping_province=safe_str(row.get('Shipping Province', '') or row.get('Shipping Province Name', '')),
                shipping_country=safe_str(row.get('Shipping Country', '')),
                shipping_zip=safe_str(row.get('Shipping Zip', '')),
            ))

            if len(batch) >= BATCH_SIZE:
                total += _flush_batch(ShopifyOrder, batch)

        total += _flush_batch(ShopifyOrder, batch)
        self.stdout.write(f"  → Shopify {total}건 임포트 완료")
        return total

    @transaction.atomic
    def _import_tiktok_csv(self, file_path, clear_date):
        """TikTok All order CSV - 스트리밍 배치 임포트"""

        def detect_brand(row):
            sku = safe_str(row.get('Seller SKU', '')).upper()
            product = safe_str(row.get('Product Name', '')).lower()
            if sku.startswith('DR-') or 'dr.blet' in product or 'pooeng' in product:
                return '닥터블릿'
            if sku.startswith('CALO-') or 'calo' in product:
                return 'Calo'
            return ''

        # Pass 1: 날짜 수집
        if clear_date:
            dates = _collect_csv_dates(file_path, 'Created Time', 'Paid Time')
            if dates:
                min_d, max_d = min(dates), max(dates)
                deleted = TiktokOrder.objects.filter(
                    region='us', order_date__gte=min_d, order_date__lte=max_d
                ).delete()[0]
                self.stdout.write(f"  기존 데이터 {deleted}건 삭제 ({min_d} ~ {max_d})")
            del dates

        # Pass 2: 스트리밍 임포트
        batch = []
        total = 0
        for row in _csv_stream(file_path):
            order_date = safe_date(row.get('Created Time') or row.get('Paid Time'))
            if not order_date:
                continue

            order_id = safe_str(row.get('Order ID', ''))
            if not order_id:
                continue

            sku_subtotal = safe_decimal(row.get('SKU Subtotal After Discount'), None)
            order_amount = safe_decimal(row.get('Order Amount'), None)

            batch.append(TiktokOrder(
                region='us',
                brand=detect_brand(row),
                final_amount=sku_subtotal if sku_subtotal is not None else order_amount,
                order_date=order_date,
                cancel_date=safe_date(row.get('Cancelled Time')),
                order_id=order_id,
                order_status=safe_str(row.get('Order Status', '')),
                seller_sku=safe_str(row.get('Seller SKU', '')),
                product_name=safe_str(row.get('Product Name', '')),
                quantity=safe_int(row.get('Quantity', 0)),
                unit_price=safe_decimal(row.get('SKU Unit Original Price'), None),
                order_amount=order_amount,
                refund_amount=safe_decimal(row.get('Order Refund Amount'), None),
                shipping_state=safe_str(row.get('State', '')),
                shipping_city=safe_str(row.get('City', '')),
                shipping_country=safe_str(row.get('Country', '')),
            ))

            if len(batch) >= BATCH_SIZE:
                total += _flush_batch(TiktokOrder, batch)

        total += _flush_batch(TiktokOrder, batch)
        self.stdout.write(f"  → TikTok {total}건 임포트 완료")
        return total

    @transaction.atomic
    def _import_shopee_excel(self, file_path, filename, clear_date):
        """Shopee shop-stats Excel 임포트 (소규모 데이터)"""
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        self.stdout.write(f"  Shopee Excel 시트: {wb.sheetnames}")

        file_date = extract_date_from_filename(filename)
        brand = extract_brand_from_shopee_filename(filename)
        self.stdout.write(f"  감지된 날짜: {file_date}, 브랜드: {brand}")

        if clear_date and file_date:
            deleted = ShopeeOrder.objects.filter(
                region='cn', order_date=file_date
            ).delete()[0]
            self.stdout.write(f"  기존 데이터 {deleted}건 삭제 ({file_date})")

        order_date = file_date
        objects = []

        if 'Placed Order' in wb.sheetnames:
            ws = wb['Placed Order']
            rows_data = list(ws.rows)
            if len(rows_data) >= 2:
                header_row = rows_data[0]
                data_row = rows_data[1]
                date_str = str(data_row[0].value or '')
                if not order_date and date_str:
                    match = re.search(r'(\d{2})-(\d{2})-(\d{4})', date_str)
                    if match:
                        day, month, year = match.groups()
                        order_date = date(int(year), int(month), int(day))

                daily_sales = safe_decimal(data_row[1].value if len(data_row) > 1 else None)
                daily_orders = safe_int(data_row[3].value if len(data_row) > 3 else None)
                daily_visitors = safe_int(data_row[6].value if len(data_row) > 6 else None)
                refunded_sales = safe_decimal(data_row[11].value if len(data_row) > 11 else None)

                if order_date and (daily_sales or daily_orders):
                    objects.append(ShopeeOrder(
                        region='cn', brand=brand, final_amount=daily_sales,
                        order_date=order_date, order_id=f'DAILY-{order_date}',
                        order_status='Daily Summary',
                        product_name=f'일별 집계 (주문 {daily_orders}건, 방문자 {daily_visitors}명)',
                        quantity=daily_orders, order_amount=daily_sales,
                        refund_amount=refunded_sales, buyer_country='SG',
                    ))

        for sheet_prefix in ['Product Contribution (place', 'Product Contribution (paid']:
            target_sheet = None
            for sname in wb.sheetnames:
                if sname.startswith(sheet_prefix):
                    target_sheet = sname
                    break
            if not target_sheet:
                continue

            ws = wb[target_sheet]
            order_type = 'Placed' if 'place' in target_sheet.lower() else 'Paid'

            for idx, row_data in enumerate(ws.rows):
                if idx < 4:  # 헤더 영역 스킵
                    continue
                cells = [c.value for c in row_data]
                item_id = str(cells[0] or '') if cells else ''
                product = str(cells[1] or '') if len(cells) > 1 else ''
                if not item_id or not item_id.replace('.', '').isdigit():
                    continue
                sales = safe_decimal(cells[4] if len(cells) > 4 else None)
                units = safe_int(cells[8] if len(cells) > 8 else None)
                if order_date and product:
                    objects.append(ShopeeOrder(
                        region='cn', brand=brand, final_amount=sales,
                        order_date=order_date, order_id=item_id,
                        order_status=order_type, product_name=product,
                        quantity=units, order_amount=sales, buyer_country='SG',
                    ))

        ShopeeOrder.objects.bulk_create(objects, batch_size=BATCH_SIZE)
        self.stdout.write(f"  → Shopee {len(objects)}건 임포트 완료")
        wb.close()
        return len(objects)

    @transaction.atomic
    def _import_qoo10_excel(self, file_path, filename, clear_date):
        """Qoo10 Transaction Excel 임포트"""
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb['data']

        order_date = extract_date_from_filename(filename)
        if not order_date:
            wb.close()
            raise CommandError(f'Qoo10 파일명에서 날짜를 추출할 수 없습니다: {filename}')

        self.stdout.write(f"  날짜: {order_date}")

        if clear_date:
            deleted = Qoo10Order.objects.filter(
                region='jp', order_date=order_date
            ).delete()[0]
            self.stdout.write(f"  기존 데이터 {deleted}건 삭제 ({order_date})")

        qoo10_brand_map = {
            'nothingbetter': '낫띵베럴', 'nothingviral': '낫띵베럴',
            'drblet': '닥터블릿', 'doctorblet': '닥터블릿', 'dr.blet': '닥터블릿',
        }

        # read_only 모드: iter_rows로 스트리밍
        headers = []
        batch = []
        total = 0
        for idx, row_data in enumerate(ws.rows):
            cells = [c.value for c in row_data]
            if idx == 0:
                headers = [str(c or '') for c in cells]
                col = {h: i for i, h in enumerate(headers)}
                continue

            product_id = safe_str(cells[col.get('상품번호', 0)] if col.get('상품번호', 0) < len(cells) else None)
            if not product_id:
                continue

            brand_idx = col.get('브랜드명', 3)
            brand_raw = safe_str(cells[brand_idx] if brand_idx < len(cells) else None)
            name = brand_raw.split('/')[0].strip().lower() if brand_raw else ''
            brand = qoo10_brand_map.get(name, brand_raw.split('/')[0].strip() if brand_raw else '')

            batch.append(Qoo10Order(
                region='jp', brand=brand,
                final_amount=safe_decimal(cells[col.get('취소분반영 거래금액', 6)] if col.get('취소분반영 거래금액', 6) < len(cells) else None, None),
                order_date=order_date,
                order_id=product_id,
                order_status='Transaction',
                product_name=safe_str(cells[col.get('상품명', 2)] if col.get('상품명', 2) < len(cells) else None),
                seller_sku=safe_str(cells[col.get('판매자상품코드', 1)] if col.get('판매자상품코드', 1) < len(cells) else None),
                quantity=safe_int(cells[col.get('취소분반영 거래상품수량', 9)] if col.get('취소분반영 거래상품수량', 9) < len(cells) else None),
                order_amount=safe_decimal(cells[col.get('거래금액', 4)] if col.get('거래금액', 4) < len(cells) else None, None),
                refund_amount=safe_decimal(cells[col.get('거래취소금액', 5)] if col.get('거래취소금액', 5) < len(cells) else None, None),
            ))

            if len(batch) >= BATCH_SIZE:
                total += _flush_batch(Qoo10Order, batch)

        total += _flush_batch(Qoo10Order, batch)
        self.stdout.write(f"  → Qoo10 {total}건 임포트 완료")
        wb.close()
        return total
